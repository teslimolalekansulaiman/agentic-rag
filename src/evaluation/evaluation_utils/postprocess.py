import pandas as pd
from openpyxl import load_workbook

import pandas as pd
from openpyxl import load_workbook

def update_results_and_summary(df, df_processed, excel_path, sheet_name):
    """
    Update the Excel file in place with new RAGAS results
    without recreating or overwriting unrelated sheets.
    Compatible with pandas >= 2.1 and openpyxl >= 3.1.
    """

    print("\n📊 Final RAGAS Summary")
    print("=" * 60)
    total = len(df)
    evaluated = len(df_processed)

    print(f"Total questions: {total}")
    print(f"Evaluated:       {evaluated}")

    metrics = ["faithfulness", "answer_relevancy", "context_recall", "context_precision"]
    for m in metrics:
        vals = df.loc[df_processed.index, m].dropna()
        if len(vals) > 0:
            print(f"  {m:20s}: {vals.mean():.4f} ± {vals.std():.4f}")

    # ✅ Safely update or create Excel file
    try:
        # Load workbook first
        book = load_workbook(excel_path)

        with pd.ExcelWriter(
            excel_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="overlay",
        ) as writer:
            # openpyxl handles sheet references internally now
            writer._book = book  # internal hook; allowed in pandas 2.x
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"\n✅ Successfully updated existing Excel file: {excel_path}")

    except FileNotFoundError:
        df.to_excel(excel_path, sheet_name=sheet_name, index=False)
        print(f"\n📁 File not found, created new Excel file: {excel_path}")


def print_metrics(metrics: dict):
    """
    Nicely formats and prints metric results after each evaluation.
    """
    print(f"    ✅ Results:")
    try:
        print(f"       Faithfulness:      {metrics['faithfulness']:.4f}")
        print(f"       Answer Relevancy:  {metrics['answer_relevancy']:.4f}")
        print(f"       Context Recall:    {metrics['context_recall']:.4f}")
        print(f"       Context Precision: {metrics['context_precision']:.4f}")
    except Exception:
        print("       ⚠️ Some metrics missing or invalid values")


def process_single_question(df, idx, row, ctx):
    """
    Handles full evaluation flow for one question using the given context.
    """
    q = str(row["question"])
    gt = str(row["ground_truth"])

    # 🧠 Generate fresh answer
    print("    🤖 Generating answer via RAG pipeline...")
    crew_result = ctx.run_crew_pipeline(q)
    answer = crew_result["answer"]
    df.at[idx, "answer"] = answer

    # ⚖️ Evaluate metrics
    print("    ⚖️ Evaluating metrics...")
    metrics = ctx.evaluate_single_question(q, answer, gt, ctx.llm, ctx.embeddings)

    # Store metrics
    for k, v in metrics.items():
        df.at[idx, k] = v

    # 💾 Save progress
    df.to_excel(ctx.excel_path, sheet_name=ctx.sheet_name, index=False)

    # 🧾 Print metrics
    ctx.print_metrics(metrics)
